
# ========================= Import all libraries ================================= #
from tkinter import * 
from customtkinter import *
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from datetime import date
from tkcalendar import *
from openpyxl import Workbook, load_workbook
import openpyxl, xlrd
import pathlib
from tkinter import messagebox

# ===========================Create the tkt window and set the appearance=========================== #
window = ctk.CTk()
window.title('STUDENT REGISTRATION FORM')
ctk.set_appearance_mode('dark')


# ==========================Open a new excel file with python coding================================ #
file =pathlib.Path("Register.xlsx")
if file.exists() :
    pass

else :
    file = Workbook()
    sheet = file.active

    # =================== making excel table headings ==================== #
    heading = ["ADMITION NUM", "DATE OF REG", "CONTACT", "STUDENT NAME", "DATE OF BIRTH", "GENDER", 'UNIVERSITY',"HEIGHER SCHOOL",'INSTITUTE','COURSE TO BE FOLLOWED']
                     
    sheet.append(heading)
    
    file.save("Register.xlsx")

# =================== Admition number counting with clicking the submit button ======================= #  
counting = 0  
def submit() :
        
    global counting
    counting += 1
    admition.set(f'GCE/OL/2023{counting}')

    # =================== Getting data from the first frame fields ============================== #
    add_01 = admition.get()
    add_02 = rdate_entry.get()
    add_03 = contact_entry.get()
    add_04 = st_name_entry.get()
    add_05 = radio.get()
    add_11 = birth_entry.get()

    # ============================ Getting the data from check box ===================== #

    if add_02 and add_03 and add_04 and add_05 and add_11 :

        add_06 = check_btn_1.get()
        add_07 = check_btn_2.get()
        add_08 = check_btn_3.get()
      
        # ================== error message if not complete the check boxes data ===================
        if add_06 == " " and add_07 == ' ' and add_08 == ' ':

            messagebox.showerror(title=" required", message="You must select at least one checkbox")

        else :
            
            # ================= get data from the combobox ============================= #
            add_10 = course_entry.get()

            #  ================== error message for the combo box =========================
            if add_10 == "Select One Of Follwing" :
                messagebox.showerror(title= "Error", message='You nust select atleast on course to be followed')

            # ================ submit all data to excell file if complete all the requierment ==============
            else :
                add_09 = terms_check_box.get()

                if add_09 == "Accepted" :
                     file = openpyxl.load_workbook("Register.xlsx")
                     sheet = file.active

                    # ================= adding data ow by row in excel file =============================
                     sheet.append([add_01,add_02,add_03,add_04,add_11, add_05,add_06,add_07,add_08,add_10])

                     file.save('Register.xlsx')

                else :
                    messagebox.showwarning(title = "Error", message = "You have not accepted the terms !")            
    else :
        messagebox.showerror(title="Required", message=" You must complete all the required details! ")

# ==============creating the main frame =============================== #
main_frame = ctk.CTkFrame(window, border_width= 1,corner_radius= 10)
main_frame.pack(pady = 75, padx = 100)
# ================== create a frame for the heading lable ========================= #
heading_frame = ctk.CTkFrame(main_frame, border_width= 1, corner_radius= 8)
heading_frame.grid(row= 0, column= 0, pady = (20,10), padx = 20)

# ==================== Heading lable place in the heading frame =================== #
heading = ctk.CTkLabel(heading_frame, text= 'PLEASE COMPLETE THE FOLLOWING DETAILS')
heading.grid(row = 0, column = 0, pady =5, padx = 40)

# ================== Make new frame to insert the entries ==========================
first_farme = ctk.CTkFrame(main_frame, border_width= 1, corner_radius= 8)
first_farme.grid(row= 1, column= 0, pady = (20,10), padx = 20)

# ========= Create admition entry with auto counting and with reset button ================= #
admition_num = ctk.CTkLabel(first_farme, text = 'ADMITION NUMBER')
admition_num.grid(row = 0, column = 0 , pady = (20,10), padx = 10)
   
admition = ctk.StringVar()
admition_entry =ctk.CTkLabel(first_farme, textvariable= admition , width = 170,bg_color='#222222',corner_radius=12)

# ================================= coding for admition number reset ================================= #
def reset() :

    global counting
    counting = 0
    admition.set(f'GCE/OL/2023{0}')
    
admition.set(f'GCE/OL/2023{0}')
admition_entry.grid(row = 1, column = 0, padx = 20, sticky = 'w')

reset_btn = ctk.CTkButton(first_farme, width= 28,height=26, text= 'R', border_width=1, fg_color='#222222',command = reset)
reset_btn.place(x = 192, y = 69)

# ======================= Creating the date of registered entry with day auto updating ================= #
today = date.today()
currebt_date = today.strftime('%d/%m/%Y')

reg_date = ctk.CTkLabel(first_farme, text = 'DATE OF REGISTER')
reg_date.grid(row = 0, column = 1, pady = (20,10), padx = 10)
rdate_entry = ctk.CTkEntry(first_farme, width= 200)
rdate_entry.insert(0, currebt_date)
rdate_entry.grid(row = 1, column = 1, padx = 20)


# ====================== Contact entry ============================================= #
contact_num = ctk.CTkLabel(first_farme, text = 'CONTACT NUMBER')
contact_num.grid(row = 0, column = 2, pady = 20, padx = 10, columnspan = 3)
contact_var = ctk.IntVar()
contact_entry = ctk.CTkEntry(first_farme, width= 200, textvariable=contact_var )
contact_entry.grid(row = 1, column = 2, padx = 20, columnspan = 3)

# ========================= Student name entry ================================================== #
student_name = ctk.CTkLabel(first_farme, text = 'STUDENT NAME')
student_name.grid(row = 2, column = 0 , pady = 10, padx = 10)
st_name_entry = ctk.CTkEntry(first_farme, width= 200)
st_name_entry.grid(row = 3, column = 0, padx = 20, pady = (0, 20))

# ======================== select the birth entry with a calander ========================== #
def birthday () :
    calnader_frame = ctk.CTkFrame(window, width= 260, height= 235,border_width= 1)
    calnader_frame.place(x = 216, y = 260)

    calander_b = Calendar(calnader_frame, selectmode = 'day', year = 2023, month = 5, day = 14)
    calander_b.place(x = 5, y = 6)

    def select_date () :
        birth_entry.delete(0, ctk.END)
        birth_entry.insert(0,calander_b.get_date())
        calnader_frame.destroy()        

    birth_submit = ctk.CTkButton(calnader_frame,text = 'SUBMIT', width= 252, command= select_date)
    birth_submit.place(x = 4 , y = 200)

date_of_birth = ctk.CTkLabel(first_farme, text = 'DATE OF BIRTH')
date_of_birth.grid(row = 2, column = 1, pady = 10, padx = 10)
birth_entry = ctk.CTkEntry(first_farme, width= 170)
birth_entry.insert(0, " ")
birth_entry.grid(row = 3, column = 1, padx = 20, pady = (0, 20), sticky = 'w')

cal_btn = ctk.CTkButton(first_farme, width= 28,height=26, text= '....', border_width=1, fg_color='#222222', command=birthday)
cal_btn.place(x = 431, y = 145)


# ============================ Select gender with radio buttons ========================= #
gender = ctk.CTkLabel(first_farme, text = 'GENDER')
gender.grid(row = 2, column = 2, pady = 10, padx = 10, columnspan = 3)

radio = StringVar()
male_button = ctk.CTkRadioButton(first_farme, text= "MALE", variable= radio, value= "Male")
male_button.grid(row = 3, column = 2, padx = (30,5), pady = (0,20))
fmale_button = ctk.CTkRadioButton(first_farme, text= "FEMALE", variable= radio, value= "Female")
fmale_button.grid(row = 3, column = 3, padx = 5, pady = (0,20))

# ========================= Create the second frame ================================== #
second_frame = ctk.CTkFrame(main_frame, border_width= 1, corner_radius= 8, width=660)
second_frame.grid(row= 2, column= 0, pady = (0,10), padx = 20)

# ====================== Select the education level with check box ========================== #
check_lbl_heading = ctk.CTkLabel(second_frame, text= "EDUCATIONAL QULIFICATION", width=165)
check_lbl_heading.grid(row = 0, column = 0, pady = (20, 0), padx = 20, columnspan = 1)

check_box_1 = tk.StringVar(value= " ")
check_btn_1 = ctk.CTkCheckBox(second_frame, text= 'University', variable = check_box_1, onvalue= "University", offvalue= ' ')
check_btn_1.deselect()
check_btn_1.grid(row = 1, column = 0, pady = (10,0), sticky = 'w',  padx = 40)
check_box_2 = tk.StringVar(value= " ")
check_btn_2 = ctk.CTkCheckBox(second_frame, text= 'Hihger School', variable = check_box_2, onvalue= "Heigher School", offvalue= " ")
check_btn_2.deselect()
check_btn_2.grid(row = 2, column = 0, pady = (10,20), sticky = 'w',  padx = 40)
check_box_3 = tk.StringVar(value = " ")
check_btn_3 = ctk.CTkCheckBox(second_frame, text= 'Institute', variable = check_box_3, onvalue= "Intitute", offvalue= " ")
check_btn_3.deselect()
check_btn_3.grid(row = 1, column = 1, pady = (10,0), sticky = 'w', padx = 40)

# ====================== Select the course with combo-box ================= #
course_str = tk.StringVar(value='Select One Of Follwing')
course_label = ctk.CTkLabel(second_frame, text= "COURSE TO BE FOLLOWED", width= 290)
course_label.grid(row = 0, column = 2, pady = (20, 0), padx = 20)
course_entry = ctk.CTkComboBox(second_frame, width= 200, values=['Python Certificate','Python Diploma', 'Python in Masters'],variable= course_str)
course_entry.grid(row = 1, column = 2)

# ======================= terms & condition with check box ======================== #
accept_var = tk.StringVar(value = 'Not Accepted')
third_frame = ctk.CTkFrame(main_frame, border_width= 1, corner_radius= 8)
third_frame.grid(row= 3, column= 0, pady = (0,10), padx = 20)

terms_check_box = ctk.CTkCheckBox(third_frame, text= 'I accept the Terms and Condition', width= 645, variable= accept_var, onvalue= "Accepted", offvalue = "Not Accepted")
terms_check_box.grid(row = 0, column = 0, pady = (20,20), sticky = 'w', padx = 40)

# =========================== Submit button =================================== #

submit_button = ctk.CTkButton(main_frame, text= 'SUBMIT', width= 725, command= submit)
submit_button.grid(row = 4, column = 0, pady = (0, 20), padx = 15)



window.mainloop()
